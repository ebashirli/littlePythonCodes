from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\Ilkin\Desktop\RFI (Request) Registration M5 elvin.xlsx")
rfi_list = wb['rfi_list']

itinerary = wb['itinerary']

itinerary_kms = list(itinerary.values)[48]

for i in range(1, 175):
    rfi_val = list(rfi_list.values)[i]
    from_col = itinerary_kms.index(rfi_val[1]) + 1
    to_col = itinerary_kms.index(rfi_val[2]) + 1

    
    if rfi_val[3] == 'LHS - RHS':
       row_no = 39 - 2 * rfi_val[5]
       for j in range(from_col, to_col):
          cell_val = itinerary.cell(row=row_no, column=j).value
          if cell_val == None:
              itinerary.cell(row=row_no, column=j).value = rfi_val[0]
          else:
             print(cell_val, rfi_val[0])
               
       row_no = 59 + 2 * rfi_val[5]
       for j in range(from_col, to_col):
          cell_val = itinerary.cell(row=row_no, column=j).value
          if cell_val == None:
              itinerary.cell(row=row_no, column=j).value = rfi_val[0]
          else:
             print(cell_val, rfi_val[0])

    elif rfi_val[3] == 'LHS':
        row_no = 39 - 2 * rfi_val[5]
        for j in range(from_col, to_col):
           cell_val = itinerary.cell(row=row_no, column=j).value
           if cell_val == None:
              itinerary.cell(row=row_no, column=j).value = rfi_val[0]
           else:
              print(cell_val, rfi_val[0])
    elif rfi_val[3] == 'RHS':
        row_no = 59 + 2 * rfi_val[5]
        for j in range(from_col, to_col):
            cell_val = itinerary.cell(row=row_no, column=j).value
            if cell_val == None:
               itinerary.cell(row=row_no, column=j).value = rfi_val[0]
            else:
               print(cell_val, rfi_val[0])

wb.save(r"C:\Users\Ilkin\Desktop\RFI (Request) Registration M5 elvin.xlsx")

