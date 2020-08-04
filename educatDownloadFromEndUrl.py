# import xlrd
# import xlwt
# from os import listdir

from webbrowser import get
from openpyxl import load_workbook

filePath = r"D:\Shared files\Education_com\Lists - English.xlsx"

wb = load_workbook(filename = filePath)
ws = wb['Workshhets-Separate']


def downloader(url):
    chrome_path = 'C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s'
    get(chrome_path).open(url)

for i in range(1, 10): 
    try:
        downloader('https://www.education.com/download/worksheet/' + str(int(ws['E' + str(i)].value)) + '/answer/' + ws['F' + str(i)].value)
        # ws['J' + str(i)] = 'ok'

    except:
        continue


wb.save(filePath)

